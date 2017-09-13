import re
import xlrd, xlwt
import openpyxl as excel
import codecs

subscribersCount = 2
index = 0

# Open excel file
# TODO Make excel file not rewriting, but adding. Connect to database maybe
wb = excel.load_workbook(filename='Test_book.xlsx')
sheet = wb['Test']

# # Add titles line in excel
# titles = ['Name','Fed okrug','Region', 'Organisation', 'Dolzhnost', 'Telephone',
#                      'E-mail']
# for i in range(0,6):
#     sheet.cell(row=1,column=i).value = titles[i]

with codecs.open('subscribers', 'r', 'utf-8') as f:
    lines = f.readlines()
    for line in lines:
        if line in ['\n', '\r\n']:
            subscribersCount += 1
            index = 0
            continue
        else:
            index += 1
        # print str(subscribersCount) + ' ' + str(index)

        line = line.replace('\n','').replace('\r','')
        sheet.cell(row=subscribersCount, column=index).value = re.sub('^.*: ', '', line)

wb.save('Test_book.xlsx')



# listOfSubscribers = []
# currentSubscriber = []
# currentSubscriber1 = {'Name': '', 'Fed okrug': '', 'Region': '', 'Organisation': '', 'Dolzhnost': '', 'Telephone': '',
#                      'E-mail': ''}


# currentSubscriber['Name'] = re.sub('^.*: ', '', line)
# currentSubscriber['Fed okrug'] = re.sub('^.*: ', '', line)
# currentSubscriber['Region'] = re.sub('^.*: ', '', line)
# currentSubscriber['Organisation'] = re.sub('^.*: ', '', line)
# currentSubscriber['Dolzhnost'] = re.sub('^.*: ', '', line)
# currentSubscriber['Telephone'] = re.sub('^.*: ', '', line)
# currentSubscriber['E-mail'] = re.sub('^.*: ', '', line)
# print re.sub('^.*: ', '', line)
# currentSubscriber.append(line)
# print currentSubscriber

# for x in listOfSubscribers:
#     print x
